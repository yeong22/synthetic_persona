"""
Privacy Calculus GT 전처리 스크립트.
- data/privacy_data_gt/results.xlsx 에서 설문 응답 로드
- data/privacy_data_gt/q_list.xlsx 에서 질문-construct 매핑
- 38개 문항을 6개 construct (PC1-PC6)로 집계
- 국가별 필터링 및 분포 JSON 저장

Construct 매핑 (q_list의 날짜 코드 → construct):
  PC1 (perceived_benefit):     01-04, 01-05, 01-06  — 개인화 서비스 혜택
  PC2 (perceived_risk):        02-09, 02-10, 02-11  — 데이터 오용/통제 우려
  PC3 (trust_provider):        02-05, 02-06, 02-07  — 서비스 제공자 신뢰
  PC4 (info_sensitivity):      02-12, 02-13, 02-14, 02-15 — 개인정보 민감성
  PC5 (sharing_intention):     01-01, 01-02, 01-03  — 정보 공유 의향
  PC6 (privacy_concern):       02-08, 02-10, 02-11  — 프라이버시 우려
"""

import json
from pathlib import Path

import numpy as np
import pandas as pd
import openpyxl

DATA_DIR = Path(__file__).resolve().parent
SRC_DIR = DATA_DIR / "privacy_data_gt"
OUT_DIR = DATA_DIR / "privacy_gt"
OUT_DIR.mkdir(exist_ok=True)

# ---------------------------------------------------------------------------
# Construct ↔ 문항 매핑 (date code → construct)
# ---------------------------------------------------------------------------
CONSTRUCT_MAP = {
    "PC1": {
        "name": "perceived_benefit",
        "items": ["2025-01-04", "2025-01-05", "2025-01-06"],
        "description": "Perceived benefit of sharing personal data",
    },
    "PC2": {
        "name": "perceived_risk",
        "items": ["2025-02-09", "2025-02-10", "2025-02-11"],
        "description": "Perceived risk of data misuse / lack of control",
    },
    "PC3": {
        "name": "trust_provider",
        "items": ["2025-02-05", "2025-02-06", "2025-02-07"],
        "description": "Trust in the service provider",
    },
    "PC4": {
        "name": "info_sensitivity",
        "items": ["2025-02-12", "2025-02-13", "2025-02-14", "2025-02-15"],
        "description": "Personal information sensitivity",
    },
    "PC5": {
        "name": "sharing_intention",
        "items": ["2025-01-01", "2025-01-02", "2025-01-03"],
        "description": "Willingness to share personal data",
    },
    "PC6": {
        "name": "privacy_concern",
        "items": ["2025-02-08", "2025-02-10", "2025-02-11"],
        "description": "General privacy concern when using AI services",
    },
}

# 국가명 정규화 (원본 데이터의 spelling 변형 통합)
COUNTRY_NORMALIZE = {
    "south africa": "South Africa",
    "south african": "South Africa",
    "south arica": "South Africa",
    "italy": "Italy",
    "united kingdom": "United Kingdom",
    "uk": "United Kingdom",
    "england": "United Kingdom",
    "germany": "Germany",
    "poland": "Poland",
    "spain": "Spain",
    "portugal": "Portugal",
    "canada": "Canada",
    "kenya": "Kenya",
    "mexico": "Mexico",
    "méxico": "Mexico",
    "czech republic": "Czech Republic",
    "czechia": "Czech Republic",
    "greece": "Greece",
    "chile": "Chile",
    "brazil": "Brazil",
    "france": "France",
    "turkey": "Turkey",
    "indonesia": "Indonesia",
    "united states": "United States",
    "india": "India",
    "finland": "Finland",
}

COUNTRY_COL = "2025-03-01"  # 국가 질문
MIN_COUNTRY_N = 10  # 최소 응답 수


def load_results() -> pd.DataFrame:
    """Load results.xlsx into DataFrame with string date-code columns."""
    wb = openpyxl.load_workbook(SRC_DIR / "results.xlsx")
    ws = wb.active

    headers = [str(c.value)[:10] for c in next(ws.iter_rows(max_row=1))]
    data = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        data.append(list(row))
    wb.close()

    df = pd.DataFrame(data, columns=headers)
    return df


def normalize_country(raw: str) -> str:
    if not isinstance(raw, str):
        return "Unknown"
    return COUNTRY_NORMALIZE.get(raw.strip().lower(), raw.strip())


def compute_construct_scores(df: pd.DataFrame) -> pd.DataFrame:
    """Compute PC1-PC6 construct scores as item means (rounded to nearest int for 1-7 scale)."""
    result = pd.DataFrame()
    result["country"] = df[COUNTRY_COL]

    for pc_code, info in CONSTRUCT_MAP.items():
        item_cols = info["items"]
        # Convert to numeric
        item_data = df[item_cols].apply(pd.to_numeric, errors="coerce")
        # Construct score = mean of items, rounded to nearest integer (keeps 1-7 scale)
        result[pc_code] = item_data.mean(axis=1).round().astype("Int64")

    return result


def main():
    print("Loading results.xlsx ...")
    df = load_results()
    print(f"  Raw shape: {df.shape}")

    # Normalize country names
    df[COUNTRY_COL] = df[COUNTRY_COL].apply(normalize_country)

    # Compute construct scores
    scores = compute_construct_scores(df)
    pc_cols = [f"PC{i}" for i in range(1, 7)]

    # Country distribution
    print(f"\n{'Country':<25} {'N':>5}")
    print("=" * 35)
    for country, count in scores["country"].value_counts().sort_values(ascending=False).items():
        marker = " ✓" if count >= MIN_COUNTRY_N else ""
        print(f"  {country:<23} {count:>5}{marker}")

    # Filter countries with sufficient data
    valid_countries = (
        scores["country"].value_counts()
        .loc[lambda x: x >= MIN_COUNTRY_N]
        .index.tolist()
    )
    print(f"\nCountries with N≥{MIN_COUNTRY_N}: {valid_countries}")

    # Build distributions and save
    distributions = {}

    print(f"\n{'Country':<25} {'Valid N':>8}  Construct score ranges")
    print("=" * 75)

    for country in sorted(valid_countries):
        cdf = scores[scores["country"] == country][pc_cols].dropna()

        distributions[country] = {}
        range_info = []
        for pc in pc_cols:
            freq = cdf[pc].value_counts().sort_index()
            distributions[country][pc] = {str(int(k)): int(v) for k, v in freq.items()}
            range_info.append(f"{pc}:{int(cdf[pc].min())}-{int(cdf[pc].max())}")

        # Save per-country CSV
        cdf.to_csv(OUT_DIR / f"{country}.csv", index=False)
        print(f"  {country:<23} {len(cdf):>8}  {', '.join(range_info)}")

    # Save distributions
    dist_path = OUT_DIR / "distributions.json"
    with open(dist_path, "w") as f:
        json.dump(distributions, f, indent=2, ensure_ascii=False)

    print(f"\n{'=' * 75}")
    print("Saved:")
    print(f"  distributions → {dist_path}")
    for country in sorted(valid_countries):
        csv_path = OUT_DIR / f"{country}.csv"
        print(f"  {country:<23} → {csv_path}")

    # Save construct mapping for reference
    mapping_path = OUT_DIR / "construct_mapping.json"
    with open(mapping_path, "w") as f:
        json.dump(CONSTRUCT_MAP, f, indent=2, ensure_ascii=False)
    print(f"  construct mapping → {mapping_path}")

    # Report countries that could be used in experiment
    print(f"\n{'=' * 75}")
    print("NOTE: experiment_config.py PRIVACY_COUNTRIES should be updated to match")
    print(f"available GT countries: {valid_countries}")


if __name__ == "__main__":
    main()
